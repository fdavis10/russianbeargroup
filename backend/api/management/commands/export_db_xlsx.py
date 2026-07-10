import os
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from uuid import UUID

from django.apps import apps
from django.core.management.base import BaseCommand
from django.db import models
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Экспортируем все таблицы проекта + пользователей админки.
EXPORT_APP_LABELS = {"api", "auth"}
EXPORT_AUTH_MODELS = {"user"}
SKIP_MODELS = {
    ("auth", "permission"),
    ("auth", "group"),
}


def serialize_value(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        if isinstance(value, datetime) and timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.isoformat(sep=" ", timespec="seconds") if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, (Decimal, UUID)):
        return str(value)
    if isinstance(value, models.Model):
        return value.pk
    if isinstance(value, (list, tuple, dict, set)):
        return str(value)
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, bool):
        return value
    return value


def sheet_title(app_label: str, model_name: str) -> str:
    title = f"{app_label}_{model_name}"[:31]
    return title


class Command(BaseCommand):
    help = "Экспорт данных БД в XLSX (лист = таблица, заголовки = имена полей модели)"

    def add_arguments(self, parser):
        parser.add_argument(
            "--output-dir",
            default=os.getenv("BACKUP_DIR", "backups"),
            help="Каталог для сохранения XLSX",
        )
        parser.add_argument(
            "--retention-days",
            type=int,
            default=int(os.getenv("BACKUP_RETENTION_DAYS", "30")),
            help="Сколько дней хранить старые бэкапы",
        )

    def handle(self, *args, **options):
        output_dir = Path(options["output_dir"])
        output_dir.mkdir(parents=True, exist_ok=True)

        today = timezone.localdate().isoformat()
        filepath = output_dir / f"backup_{today}.xlsx"
        latest_path = output_dir / "backup_latest.xlsx"

        workbook = Workbook()
        workbook.remove(workbook.active)

        exported_rows = 0
        exported_models = 0

        for model in apps.get_models():
            if model._meta.proxy or not model._meta.managed:
                continue

            app_label = model._meta.app_label
            model_name = model._meta.model_name
            model_key = (app_label, model_name)

            if app_label not in EXPORT_APP_LABELS:
                continue
            if app_label == "auth" and model_name not in EXPORT_AUTH_MODELS:
                continue
            if model_key in SKIP_MODELS:
                continue

            fields = list(model._meta.fields)
            headers = [field.name for field in fields]

            sheet = workbook.create_sheet(title=sheet_title(app_label, model_name))
            sheet.append(headers)

            queryset = model._default_manager.all().order_by("pk")
            for obj in queryset.iterator():
                row = [serialize_value(getattr(obj, field.name)) for field in fields]
                sheet.append(row)
                exported_rows += 1

            for column_index, header in enumerate(headers, start=1):
                column_letter = get_column_letter(column_index)
                sheet.column_dimensions[column_letter].width = min(max(len(header) + 2, 12), 40)

            exported_models += 1
            self.stdout.write(
                f"  {app_label}.{model._meta.object_name}: {queryset.count()} записей"
            )

        if exported_models == 0:
            self.stderr.write(self.style.WARNING("Нет моделей для экспорта."))
            return

        workbook.save(filepath)
        workbook.save(latest_path)

        removed = self._cleanup_old_backups(output_dir, options["retention_days"])

        self.stdout.write(
            self.style.SUCCESS(
                f"Бэкап сохранён: {filepath} "
                f"({exported_models} таблиц, {exported_rows} строк). "
                f"Удалено старых файлов: {removed}."
            )
        )

    def _cleanup_old_backups(self, output_dir: Path, retention_days: int) -> int:
        if retention_days <= 0:
            return 0

        cutoff = timezone.localdate() - timedelta(days=retention_days)
        removed = 0

        for path in output_dir.glob("backup_????-??-??.xlsx"):
            try:
                file_date = date.fromisoformat(path.stem.replace("backup_", ""))
            except ValueError:
                continue
            if file_date < cutoff:
                path.unlink(missing_ok=True)
                removed += 1

        return removed
